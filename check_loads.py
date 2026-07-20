#!/usr/bin/env python3
"""Check Harlo load-confirmation status for every load in a TC export.

Reads unique load numbers from the WBSHPGRP column of the input workbook,
POSTs each to https://harlo.gambitco.io/api/load-confirmation/initiate using
the session captured by capture_auth.py, and writes a copy of the workbook
with a "Load Status" tab (<input stem>_results.xlsx).

Responses are appended to <input stem>_results.cache.jsonl as they arrive,
so an interrupted run resumes where it left off; pass --fresh to re-check
everything.

Usage:
    python check_loads.py Load_Confirmation_2026-07-20.xlsx [options]
"""

import argparse
import json
import re
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor
from datetime import datetime, timezone
from pathlib import Path
from shutil import copyfile

import requests
from openpyxl import load_workbook
from openpyxl.styles import Font

BASE_URL = "https://harlo.gambitco.io"
API_URL = f"{BASE_URL}/api/load-confirmation/initiate"
AUTH_FILE = Path(__file__).parent / ".auth" / "harlo.json"
LOAD_COLUMN = "WBSHPGRP"
RESULT_SHEET = "Load Status"
_AUTH_LOCK = threading.Lock()


def extract_loads(xlsx_path: Path, sheet: str | None) -> list[str]:
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    if LOAD_COLUMN not in header:
        sys.exit(f"Column {LOAD_COLUMN!r} not found in sheet {ws.title!r}. Headers: {header}")
    col = header.index(LOAD_COLUMN)
    seen: dict[str, None] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        value = row[col]
        if value not in (None, ""):
            seen[str(value).strip()] = None
    wb.close()
    return list(seen)


def build_session() -> requests.Session:
    if not AUTH_FILE.exists():
        sys.exit(f"No cached session at {AUTH_FILE}. Run: python capture_auth.py")
    state = json.loads(AUTH_FILE.read_text())
    session = requests.Session()
    session.headers.update(
        {
            "Content-Type": "application/json",
            "Origin": BASE_URL,
            "Referer": f"{BASE_URL}/apps/load-confirmation",
        }
    )
    for cookie in state.get("cookies", []):
        session.cookies.set(
            cookie["name"], cookie["value"], domain=cookie.get("domain"), path=cookie.get("path", "/")
        )
    # If the app keeps Cognito-style tokens in localStorage, hold onto them as
    # a fallback Authorization header for when cookies alone get a 401/403.
    tokens = {}
    for origin in state.get("origins", []):
        for item in origin.get("localStorage", []):
            for kind in ("idToken", "accessToken"):
                if item["name"].endswith(f".{kind}"):
                    tokens[kind] = item["value"]
    session.fallback_tokens = tokens  # type: ignore[attr-defined]
    return session


def check_one(session: requests.Session, load_number: str, retries: int = 3) -> dict:
    last_error = None
    for attempt in range(retries):
        try:
            resp = session.post(API_URL, json={"loadNumber": load_number}, timeout=60)
            if resp.status_code in (401, 403):
                tokens = getattr(session, "fallback_tokens", {})
                with _AUTH_LOCK:
                    if tokens and "Authorization" not in session.headers:
                        token = tokens.get("idToken") or tokens.get("accessToken")
                        session.headers["Authorization"] = f"Bearer {token}"
                        continue
                sys.exit(
                    f"Got HTTP {resp.status_code} from the API — session expired. Run: python capture_auth.py"
                )
            resp.raise_for_status()
            return resp.json()
        except (requests.RequestException, ValueError) as exc:
            last_error = exc
            time.sleep(2**attempt)
    return {"error": str(last_error)}


def derive_status(response: dict) -> tuple[str, str]:
    """Return (status, first exception) for one API response.

    Statuses beyond Ready to Confirm are one "Exception" bucket for now;
    once a full run shows the range of exception texts they can be grouped
    into categories here.
    """
    if "error" in response:
        return "ERROR", response["error"]
    if not response.get("success"):
        return "ERROR", json.dumps(response)[:500]
    result = response.get("result") or {}
    exceptions = result.get("exceptions") or []
    if exceptions:
        return "Exception", exceptions[0]
    if result.get("halted"):
        halts = [i.get("text", "").strip() for i in result.get("reasoning") or [] if i.get("type") == "halt"]
        return "Exception", halts[0] if halts else "halted with no exception text"
    if not response.get("ready"):
        return "Not Ready", ""
    return "Ready to Confirm", ""


def categorize(exception_text: str) -> str:
    """Collapse an exception into a category by masking every number, so
    'Avg cube/lift 13.33 below min (20)' and 'Avg cube/lift 9.5 below min (20)'
    group as 'Avg cube/lift # below min (#)'."""
    return re.sub(r"\s+", " ", re.sub(r"\d+(?:\.\d+)?", "#", exception_text)).strip()


def write_results(input_path: Path, output_path: Path, results: dict[str, dict]) -> dict[str, int]:
    copyfile(input_path, output_path)
    wb = load_workbook(output_path)
    if RESULT_SHEET in wb.sheetnames:
        del wb[RESULT_SHEET]
    ws = wb.create_sheet(RESULT_SHEET)
    headers = ["Load #", "Status", "Category", "First Exception", "Checked At"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    counts: dict[str, int] = {}
    for load_number, entry in results.items():
        status, details = derive_status(entry["response"])
        counts[status] = counts.get(status, 0) + 1
        category = categorize(details) if status == "Exception" else ""
        ws.append([load_number, status, category, details, entry["checked_at"]])
    for column, width in zip(ws.columns, (12, 22, 44, 80, 24)):
        ws.column_dimensions[column[0].column_letter].width = width
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)
    return counts


def load_cache(cache_path: Path) -> dict[str, dict]:
    results: dict[str, dict] = {}
    if cache_path.exists():
        for line in cache_path.read_text().splitlines():
            if line.strip():
                entry = json.loads(line)
                results[entry["loadNumber"]] = entry
    return results


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("input", type=Path, help="TC export workbook (.xlsx)")
    parser.add_argument("--sheet", help="sheet name holding the data (default: first sheet)")
    parser.add_argument("--fresh", action="store_true", help="ignore the cache and re-check every load")
    parser.add_argument("--delay", type=float, default=0.3, help="seconds each worker pauses between calls (default 0.3)")
    parser.add_argument("--workers", type=int, default=4, help="parallel API calls (default 4)")
    parser.add_argument("--limit", type=int, help="only check the first N loads (for testing)")
    args = parser.parse_args()

    if not args.input.exists():
        sys.exit(f"Input file not found: {args.input}")
    output_path = args.input.with_name(f"{args.input.stem}_results.xlsx")
    cache_path = args.input.with_name(f"{args.input.stem}_results.cache.jsonl")

    loads = extract_loads(args.input, args.sheet)
    if args.limit:
        loads = loads[: args.limit]
    print(f"{len(loads)} unique {LOAD_COLUMN} load numbers in {args.input.name}")

    if args.fresh:
        cache_path.unlink(missing_ok=True)
    results = load_cache(cache_path)
    pending = [ln for ln in loads if ln not in results]
    if len(results) and pending:
        print(f"Resuming: {len(loads) - len(pending)} already cached, {len(pending)} to check")

    if pending:
        session = build_session()
        record_lock = threading.Lock()
        done = 0
        with cache_path.open("a") as cache:

            def check_and_record(load_number: str) -> None:
                nonlocal done
                entry = {
                    "loadNumber": load_number,
                    "checked_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
                    "response": check_one(session, load_number),
                }
                status, _ = derive_status(entry["response"])
                with record_lock:
                    results[load_number] = entry
                    cache.write(json.dumps(entry) + "\n")
                    cache.flush()
                    done += 1
                    print(f"[{done}/{len(pending)}] {load_number}: {status}")
                if args.delay:
                    time.sleep(args.delay)

            with ThreadPoolExecutor(max_workers=args.workers) as pool:
                for future in [pool.submit(check_and_record, ln) for ln in pending]:
                    future.result()

    counts = write_results(args.input, output_path, {ln: results[ln] for ln in loads})
    print(f"\nWrote {output_path.name} ({RESULT_SHEET!r} tab)")
    for status, count in sorted(counts.items(), key=lambda kv: -kv[1]):
        print(f"  {count:>5}  {status}")

    categories: dict[str, int] = {}
    for load_number in loads:
        status, details = derive_status(results[load_number]["response"])
        if status == "Exception" and details:
            category = categorize(details)
            categories[category] = categories.get(category, 0) + 1
    if categories:
        print("\nException categories:")
        for category, count in sorted(categories.items(), key=lambda kv: -kv[1]):
            print(f"  {count:>5}  {category}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
